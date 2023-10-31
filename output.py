import openpyxl
import json
import contextlib

workbook = openpyxl.load_workbook("1.3 角色基础属性.xlsx")

with contextlib.closing(workbook):
    sheet = workbook["角色基本属性"]

    characters = []

    for i in sheet.iter_rows(min_col=1):
        characters = {"name": i[0].value, "mt": i[2].value, "attack": i[3].value,
                     "defensive": i[4].value, "health": i[5].value,
                     "speed": i[6].value, "aggro": i[7].value
                      }
        characters.append(characters)

    characters = characters[1:]

    with open("data/characters.json", "w", encoding="utf-8") as f:
        json.dump(characters, f, ensure_ascii=False, indent=4)

workbook = openpyxl.load_workbook("1.3 光锥基本属性.xlsx")

with contextlib.closing(workbook):
    sheet = workbook["Sheet1"]

    attackers = []
    for i in sheet.iter_rows(min_col=1):
        attack = {"name": i[0].value, "mt": i[9].value, "health": i[1].value, "attack": i[2].value,
                  "defensive": i[3].value}
        attackers.append(attack)

    attackers = attackers[1:]

    with open("data/attackers.json", "w", encoding="utf-8") as f:
        json.dump(attackers, f, ensure_ascii=False, indent=4)


with contextlib.closing(workbook):
    sheet = workbook['11']